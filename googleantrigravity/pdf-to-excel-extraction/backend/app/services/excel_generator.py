"""
Excel file generation service with professional formatting.
"""

from pathlib import Path
from typing import Dict, Any, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from loguru import logger

from app.core.config import settings


class ExcelGenerator:
    """Generates professionally formatted Excel files from DataFrames."""
    
    def __init__(self):
        self.output_dir = Path(settings.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_excel(
        self,
        df: pd.DataFrame,
        filename: str,
        sheet_name: str = "Extracted Data",
        include_summary: bool = True,
        summary_data: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Generate a formatted Excel file from DataFrame.
        
        Args:
            df: DataFrame to export
            filename: Output filename (without extension)
            sheet_name: Name of the main data sheet
            include_summary: Whether to include a summary sheet
            summary_data: Optional summary statistics
        
        Returns:
            Path to generated Excel file
        """
        try:
            logger.info(f"Generating Excel file: {filename}")
            
            # Clean filename
            filename = self._clean_filename(filename)
            filepath = self.output_dir / f"{filename}.xlsx"
            
            # Convert date columns to string for better compatibility
            df_copy = df.copy()
            for col in df_copy.columns:
                if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                    df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Write to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write main data
                df_copy.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format main sheet
                workbook = writer.book
                worksheet = workbook[sheet_name]
                self._format_data_sheet(worksheet, df_copy)
                
                # Add summary sheet if requested
                if include_summary and summary_data:
                    self._add_summary_sheet(workbook, summary_data)
            
            file_size = filepath.stat().st_size
            logger.success(
                f"Excel file generated: {filepath} "
                f"({len(df)} rows, {len(df.columns)} columns, {file_size} bytes)"
            )
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {e}")
            raise
    
    def generate_csv(
        self,
        df: pd.DataFrame,
        filename: str
    ) -> str:
        """
        Generate a CSV file from DataFrame.
        
        Args:
            df: DataFrame to export
            filename: Output filename (without extension)
        
        Returns:
            Path to generated CSV file
        """
        try:
            logger.info(f"Generating CSV file: {filename}")
            
            # Clean filename
            filename = self._clean_filename(filename)
            filepath = self.output_dir / f"{filename}.csv"
            
            # Write to CSV
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            file_size = filepath.stat().st_size
            logger.success(f"CSV file generated: {filepath} ({file_size} bytes)")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating CSV file: {e}")
            raise
    
    def generate_multi_sheet_excel(
        self,
        dataframes: Dict[str, pd.DataFrame],
        filename: str
    ) -> str:
        """
        Generate Excel file with multiple sheets.
        
        Args:
            dataframes: Dictionary mapping sheet names to DataFrames
            filename: Output filename (without extension)
        
        Returns:
            Path to generated Excel file
        """
        try:
            logger.info(f"Generating multi-sheet Excel file: {filename}")
            
            # Clean filename
            filename = self._clean_filename(filename)
            filepath = self.output_dir / f"{filename}.xlsx"
            
            # Write to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    # Convert dates
                    df_copy = df.copy()
                    for col in df_copy.columns:
                        if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                            df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Write sheet
                    df_copy.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Format sheet
                    workbook = writer.book
                    worksheet = workbook[sheet_name]
                    self._format_data_sheet(worksheet, df_copy)
            
            file_size = filepath.stat().st_size
            logger.success(f"Multi-sheet Excel generated: {filepath} ({file_size} bytes)")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating multi-sheet Excel: {e}")
            raise
    
    def _format_data_sheet(self, worksheet, df: pd.DataFrame):
        """Apply professional formatting to data sheet."""
        try:
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Data formatting
            data_alignment = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply to all cells
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Skip header
                        cell.alignment = data_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
        except Exception as e:
            logger.warning(f"Error formatting sheet: {e}")
    
    def _add_summary_sheet(self, workbook, summary_data: Dict[str, Any]):
        """Add a summary sheet to the workbook."""
        try:
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary", 0)
            
            # Add title
            summary_sheet['A1'] = "Extraction Summary"
            summary_sheet['A1'].font = Font(size=14, bold=True)
            
            # Add timestamp
            summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Add summary data
            row = 4
            for key, value in summary_data.items():
                summary_sheet[f'A{row}'] = key
                summary_sheet[f'B{row}'] = value
                
                # Format key column
                summary_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Auto-adjust columns
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 20
            
        except Exception as e:
            logger.warning(f"Error adding summary sheet: {e}")
    
    def _clean_filename(self, filename: str) -> str:
        """Clean filename to remove invalid characters."""
        # Remove extension if present
        filename = Path(filename).stem
        
        # Replace invalid characters
        invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        
        # Add timestamp if filename is empty
        if not filename:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        return filename


# Global instance
excel_generator = ExcelGenerator()
